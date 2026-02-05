import openpyxl

wb = openpyxl.load_workbook('Context/Fact Find Data Analysis v6.1.xlsx', data_only=True, read_only=True)
ws = wb['Fact Finds']

# Find header row
header_row = None
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    if row and any('entity' in str(cell).lower() if cell else False for cell in row):
        header_row = row
        break

# Find column indices
entity_col_idx = None
field_col_idx = None
api_name_col_idx = None
api_property_col_idx = None
area_col_idx = None
notes_col_idx = None
full_ff_col_idx = None

for idx, cell in enumerate(header_row):
    if cell:
        cell_lower = str(cell).lower()
        if cell_lower == 'entity':
            entity_col_idx = idx
        elif cell_lower == 'field':
            field_col_idx = idx
        elif cell_lower == 'api name':
            api_name_col_idx = idx
        elif cell_lower == 'api property name':
            api_property_col_idx = idx
        elif cell_lower == 'area':
            area_col_idx = idx
        elif cell_lower == 'notes':
            notes_col_idx = idx
        elif 'full' in cell_lower and 'ff' in cell_lower:
            full_ff_col_idx = idx

print("Column Indices:")
print(f"  Entity: {entity_col_idx}")
print(f"  Field: {field_col_idx}")
print(f"  API Name: {api_name_col_idx}")
print(f"  API Property Name: {api_property_col_idx}")
print(f"  Area: {area_col_idx}")
print(f"  Notes: {notes_col_idx}")
print(f"  Full FF: {full_ff_col_idx}")

print("\n=== EMPLOYMENT GAPS ANALYSIS ===\n")

# Categorize fields
current_api_fields = []
gaps_no_api_name = []
gaps_unknown_property = []
gaps_missing_from_api = []
income_related = []

for i, data_row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    if not data_row or len(data_row) <= max(entity_col_idx or 0, field_col_idx or 0):
        continue

    entity_value = str(data_row[entity_col_idx]) if entity_col_idx and data_row[entity_col_idx] else ''
    area_value = str(data_row[area_col_idx]) if area_col_idx and len(data_row) > area_col_idx and data_row[area_col_idx] else ''

    if 'employ' not in entity_value.lower() and 'employ' not in area_value.lower():
        continue

    field = str(data_row[field_col_idx]) if field_col_idx and data_row[field_col_idx] else ''
    api_name = str(data_row[api_name_col_idx]) if api_name_col_idx and len(data_row) > api_name_col_idx and data_row[api_name_col_idx] else ''
    api_property = str(data_row[api_property_col_idx]) if api_property_col_idx and len(data_row) > api_property_col_idx and data_row[api_property_col_idx] else ''
    notes = str(data_row[notes_col_idx]) if notes_col_idx and len(data_row) > notes_col_idx and data_row[notes_col_idx] else ''
    full_ff = str(data_row[full_ff_col_idx]) if full_ff_col_idx and len(data_row) > full_ff_col_idx and data_row[full_ff_col_idx] else ''

    field_info = {
        'row': i,
        'field': field,
        'api_name': api_name,
        'api_property': api_property,
        'notes': notes,
        'full_ff': full_ff,
        'area': area_value,
        'entity': entity_value
    }

    # Categorize
    if 'income' in field.lower() and 'retirement' not in field.lower():
        income_related.append(field_info)
    elif api_name and 'employment' in api_name.lower() and api_property and '[' not in api_property:
        current_api_fields.append(field_info)
    elif '[n/a]' in api_name.lower() or not api_name or api_name.strip() == '':
        gaps_no_api_name.append(field_info)
    elif '[unknown]' in api_property.lower() or '[unknown]' in api_name.lower():
        gaps_unknown_property.append(field_info)
    elif 'cannot find in office' in notes.lower() or 'missing from fff' in notes.lower():
        gaps_missing_from_api.append(field_info)

print("=" * 80)
print("CURRENT API FIELDS (Already Implemented)")
print("=" * 80)
for f in current_api_fields:
    print(f"[OK] {f['field']}")
    print(f"   API Property: {f['api_property']}")
    print()

print("\n" + "=" * 80)
print("GAPS - Fields with NO API (Marked [N/A] or empty)")
print("=" * 80)
for f in gaps_no_api_name:
    print(f"[GAP] {f['field']}")
    print(f"   API Name: {f['api_name']}")
    print(f"   Notes: {f['notes'][:100] if f['notes'] else 'N/A'}")
    print()

print("\n" + "=" * 80)
print("GAPS - Fields with [Unknown] API Property")
print("=" * 80)
for f in gaps_unknown_property:
    print(f"[GAP] {f['field']}")
    print(f"   API Name: {f['api_name']}")
    print(f"   API Property: {f['api_property']}")
    print(f"   Notes: {f['notes'][:100] if f['notes'] else 'N/A'}")
    print()

print("\n" + "=" * 80)
print("GAPS - Fields Missing from Office/API (per Notes)")
print("=" * 80)
for f in gaps_missing_from_api:
    print(f"[GAP] {f['field']}")
    print(f"   API Name: {f['api_name']}")
    print(f"   Notes: {f['notes'][:150]}")
    print()

print("\n" + "=" * 80)
print("INCOME RELATED (Should be in Income API, not Employment)")
print("=" * 80)
for f in income_related:
    print(f"[INCOME] {f['field']}")
    print(f"   API Name: {f['api_name']}")
    print(f"   Notes: Reference TDetailedincomebreakdown / Income API")
    print()

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Current API Fields: {len(current_api_fields)}")
print(f"Gaps - No API: {len(gaps_no_api_name)}")
print(f"Gaps - Unknown Property: {len(gaps_unknown_property)}")
print(f"Gaps - Missing (per notes): {len(gaps_missing_from_api)}")
print(f"Income Related (separate API): {len(income_related)}")
print(f"TOTAL GAPS FOR EMPLOYMENT API: {len(gaps_no_api_name) + len(gaps_unknown_property) + len(gaps_missing_from_api)}")
