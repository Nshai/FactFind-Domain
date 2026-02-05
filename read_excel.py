import openpyxl
import sys

try:
    print('Loading workbook...')
    wb = openpyxl.load_workbook('Context/Fact Find Data Analysis v6.0.xlsx', data_only=True, read_only=True)
    print(f'\nAvailable sheets ({len(wb.sheetnames)}):')
    for i, name in enumerate(wb.sheetnames, 1):
        print(f'  {i}. {name}')

    # Try to find and read Entity or Employment related sheet
    found_entity_sheet = False
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Check first few rows for "Entity" column
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            if row and any('entity' in str(cell).lower() if cell else False for cell in row):
                print(f'\n=== Found "Entity" in sheet: {sheet_name} (Row {row_idx}) ===')
                print('Header row:', row)
                found_entity_sheet = True

                # Find Entity column index
                entity_col_idx = None
                for idx, cell in enumerate(row):
                    if cell and 'entity' in str(cell).lower():
                        entity_col_idx = idx
                        break

                # Read rows and filter for Employment
                print('\nLooking for Employment-related rows...')
                print(f'Entity column index: {entity_col_idx}')

                # Find Area column index
                area_col_idx = None
                for idx, cell in enumerate(row):
                    if cell and 'area' in str(cell).lower() and 'sub' not in str(cell).lower():
                        area_col_idx = idx
                        break

                # First, let's see all unique entity and area values
                unique_entities = set()
                unique_areas = set()
                for data_row in ws.iter_rows(min_row=row_idx+1, max_row=ws.max_row, values_only=True):
                    if data_row and entity_col_idx is not None and len(data_row) > entity_col_idx:
                        entity_value = data_row[entity_col_idx]
                        if entity_value:
                            unique_entities.add(str(entity_value))
                        if area_col_idx is not None and len(data_row) > area_col_idx:
                            area_value = data_row[area_col_idx]
                            if area_value:
                                unique_areas.add(str(area_value))

                print(f'\nAll unique Area values: {sorted(unique_areas)}')
                print(f'\nAll unique Entity values: {sorted(unique_entities)}')

                # Now filter for Employment
                print('\n=== Employment-related rows ===')
                row_count = 0
                for i, data_row in enumerate(ws.iter_rows(min_row=row_idx+1, max_row=ws.max_row, values_only=True), row_idx+1):
                    if data_row and entity_col_idx is not None and len(data_row) > entity_col_idx:
                        entity_value = str(data_row[entity_col_idx]) if data_row[entity_col_idx] else ''
                        area_value = str(data_row[area_col_idx]) if area_col_idx and len(data_row) > area_col_idx and data_row[area_col_idx] else ''

                        # Match Employment entity or Employments area
                        if ('employ' in entity_value.lower()) or ('employ' in area_value.lower()):
                            # Create formatted output
                            formatted_row = {}
                            for col_idx, cell_value in enumerate(data_row):
                                if col_idx < len(row) and row[col_idx]:
                                    formatted_row[str(row[col_idx])] = cell_value

                            print(f'\nRow {i}:')
                            for key, value in formatted_row.items():
                                if value:  # Only print non-empty values
                                    print(f'  {key}: {value}')
                            row_count += 1

                print(f'\n\nTotal Employment rows found: {row_count}')
                break

        if found_entity_sheet:
            break

    if not found_entity_sheet:
        print('\n=== No Entity column found, showing first sheet content ===')
        first_sheet = wb[wb.sheetnames[0]]
        for i, row in enumerate(first_sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            print(f'Row {i}: {row}')

except Exception as e:
    print(f'Error: {e}', file=sys.stderr)
    import traceback
    traceback.print_exc()
    sys.exit(1)
