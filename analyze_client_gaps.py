import openpyxl

wb = openpyxl.load_workbook('Context/Fact Find Data Analysis v6.1.xlsx', data_only=True, read_only=True)
ws = wb['Fact Finds']

# Find header row
header_row = None
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    if row and any('entity' in str(cell).lower() if cell else False for cell in row):
        header_row = row
        break

# Find column indices
entity_col_idx = None
field_col_idx = None
api_name_col_idx = None
api_property_col_idx = None
area_col_idx = None
notes_col_idx = None

for idx, cell in enumerate(header_row):
    if cell:
        cell_lower = str(cell).lower()
        if cell_lower == 'entity':
            entity_col_idx = idx
        elif cell_lower == 'field':
            field_col_idx = idx
        elif cell_lower == 'api name':
            api_name_col_idx = idx
        elif cell_lower == 'api property name':
            api_property_col_idx = idx
        elif cell_lower == 'area':
            area_col_idx = idx
        elif cell_lower == 'notes':
            notes_col_idx = idx

print("=== CLIENT DATA ANALYSIS FOR V3 API ===\n")

# Categories for client-related data
client_categories = {
    'personal': [],
    'contact': [],
    'address': [],
    'identification': [],
    'health': [],
    'territorial': [],
    'family': [],
    'business': [],
    'trust': [],
    'other': []
}

# Keywords to identify client-related fields
personal_keywords = ['name', 'title', 'salutation', 'birth', 'age', 'gender', 'marital', 'nationality', 'deceased']
contact_keywords = ['phone', 'email', 'mobile', 'telephone', 'contact']
address_keywords = ['address', 'postcode', 'county', 'country', 'town', 'city']
identification_keywords = ['ni number', 'passport', 'id', 'reference', 'registration']
health_keywords = ['health', 'smoker', 'medical', 'disability', 'vulnerable']
territorial_keywords = ['resident', 'domicile', 'citizenship', 'expatriate', 'territory']
family_keywords = ['dependant', 'child', 'spouse', 'partner', 'family', 'household']
business_keywords = ['company', 'business', 'vat', 'incorporation', 'director']
trust_keywords = ['trust', 'trustee', 'beneficiary', 'settlement']

for i, data_row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    if not data_row or len(data_row) <= max(entity_col_idx or 0, field_col_idx or 0):
        continue

    entity_value = str(data_row[entity_col_idx]) if entity_col_idx and data_row[entity_col_idx] else ''
    area_value = str(data_row[area_col_idx]) if area_col_idx and len(data_row) > area_col_idx and data_row[area_col_idx] else ''
    field = str(data_row[field_col_idx]) if field_col_idx and data_row[field_col_idx] else ''
    
    # Skip if not client-related
    if not any(keyword in entity_value.lower() or keyword in area_value.lower() or keyword in field.lower() 
               for keyword in ['client', 'person', 'individual', 'contact', 'personal', 'profile', 'details']):
        continue
    
    # Skip employment and income fields (separate APIs)
    if any(keyword in field.lower() for keyword in ['employment', 'employ', 'income', 'salary', 'wage']):
        continue

    api_name = str(data_row[api_name_col_idx]) if api_name_col_idx and len(data_row) > api_name_col_idx and data_row[api_name_col_idx] else ''
    api_property = str(data_row[api_property_col_idx]) if api_property_col_idx and len(data_row) > api_property_col_idx and data_row[api_property_col_idx] else ''
    notes = str(data_row[notes_col_idx]) if notes_col_idx and len(data_row) > notes_col_idx and data_row[notes_col_idx] else ''

    field_info = {
        'row': i,
        'field': field,
        'api_name': api_name,
        'api_property': api_property,
        'notes': notes,
        'area': area_value,
        'entity': entity_value
    }

    # Categorize the field
    field_lower = field.lower()
    categorized = False
    
    for keyword in personal_keywords:
        if keyword in field_lower:
            client_categories['personal'].append(field_info)
            categorized = True
            break
    
    if not categorized:
        for keyword in contact_keywords:
            if keyword in field_lower:
                client_categories['contact'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in address_keywords:
            if keyword in field_lower:
                client_categories['address'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in identification_keywords:
            if keyword in field_lower:
                client_categories['identification'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in health_keywords:
            if keyword in field_lower:
                client_categories['health'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in territorial_keywords:
            if keyword in field_lower:
                client_categories['territorial'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in family_keywords:
            if keyword in field_lower:
                client_categories['family'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in business_keywords:
            if keyword in field_lower:
                client_categories['business'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        for keyword in trust_keywords:
            if keyword in field_lower:
                client_categories['trust'].append(field_info)
                categorized = True
                break
    
    if not categorized:
        client_categories['other'].append(field_info)

# Print analysis by category
for category, fields in client_categories.items():
    if not fields:
        continue
        
    print(f"\n{'='*80}")
    print(f"{category.upper()} FIELDS ({len(fields)} fields)")
    print(f"{'='*80}")
    
    # Separate into existing vs gaps
    existing_fields = []
    gap_fields = []
    
    for field_info in fields:
        if (field_info['api_name'] and 
            '[n/a]' not in field_info['api_name'].lower() and 
            '[unknown]' not in field_info['api_name'].lower() and
            field_info['api_name'].strip() != ''):
            existing_fields.append(field_info)
        else:
            gap_fields.append(field_info)
    
    if existing_fields:
        print(f"\n--- EXISTING IN API ({len(existing_fields)} fields) ---")
        for f in existing_fields:
            print(f"[OK] {f['field']}")
            if f['api_property']:
                print(f"     API Property: {f['api_property']}")
            print()
    
    if gap_fields:
        print(f"\n--- GAPS - MISSING FROM API ({len(gap_fields)} fields) ---")
        for f in gap_fields:
            print(f"[GAP] {f['field']}")
            print(f"      API Name: {f['api_name']}")
            print(f"      Notes: {f['notes'][:100] if f['notes'] else 'N/A'}")
            print()

# Summary
print(f"\n{'='*80}")
print("SUMMARY FOR CLIENT V3 API")
print(f"{'='*80}")

total_existing = 0
total_gaps = 0

for category, fields in client_categories.items():
    if not fields:
        continue
        
    existing = len([f for f in fields if f['api_name'] and '[n/a]' not in f['api_name'].lower() and '[unknown]' not in f['api_name'].lower() and f['api_name'].strip() != ''])
    gaps = len(fields) - existing
    
    total_existing += existing
    total_gaps += gaps
    
    print(f"{category.capitalize():15}: {existing:2d} existing, {gaps:2d} gaps, {len(fields):2d} total")

print(f"\nOVERALL TOTALS: {total_existing} existing, {total_gaps} gaps, {total_existing + total_gaps} total client fields")